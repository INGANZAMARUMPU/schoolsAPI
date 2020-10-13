from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .forms import *
from .models.mark import *
from apps.Base.models import *
from apps.Prefect.models import *
from apps.Director.models import School
from django.core.paginator import Paginator
from django.contrib import messages
import openpyxl
from io import BytesIO

@login_required
def teacher_pannel_view(request, school):
	show_hidden = "hidden"
	school = get_object_or_404(School, slug=school)
	prefect = get_object_or_404(Role, url='teacher')
	attribution = get_object_or_404(Attribution, user=request.user, school=school, role=prefect)
	classes = Class.objects.filter(school=school)
	courses = Course.objects.filter(prof=request.user.profil)
	return render(request, "teacher_pannel.html", locals())

@login_required
def titular_pannel_view(request, school):
	show_hidden = "hidden"
	school = get_object_or_404(School, slug=school)
	prefect = get_object_or_404(Role, url='titular')
	attribution = get_object_or_404(Attribution, user=request.user, school=school, role=prefect)
	classes = Class.objects.filter(school=school)
	return render(request, "titular_pannel.html", locals())

@login_required
def titular_class_view(request):
	show_hidden = "hidden"
	classe = get_object_or_404(Class, titulaire = request.user.profil)
	courses = Course.objects.filter(Class=classe)
	return render(request, "titular_class.html", locals())


@login_required
def titular_course_view(request, slug):
	show_hidden = "hidden"
	classe = get_object_or_404(Class, titulaire = request.user.profil)
	course = get_object_or_404(Course,Class=classe)
	student = Student.objects.filter(Class = classe)
	students_works = StudentWork.objects.filter(student__in = student)

	return render(request, 'titular_course.html', locals())


@login_required
def add_mark_view(request, work_id, page=1):
	show_hidden = "hidden"
	work = get_object_or_404(Work, pk=work_id)
	students = Student.objects.filter(Class = work.course.Class)
	paginator = Paginator(students, 1)
	try:
		pagination = paginator.page(page)
		student = pagination.object_list[0]

		try:
			mark = Mark.objects.get(student=student, work=work)
			add_mark_view(request, work_id, page + 1)
		except:
			print
	except:
		return redirect(course_view, slug=work.course.slug)

	form = MarkForm(request.POST)
	if request.method == 'POST':
		if form.is_valid():
			marks = form.cleaned_data['marks']
			try:
				mark = Mark.objects.get(student=student, work=work)
				mark.marks = marks
				if mark.marks > mark.work.maxima:
					messages.error(request, "Mark out of maxima")
					form =  MarkForm()
				else:
					mark.save()
					page += 1
			except:
				mark = Mark(student = student, work=work)
				mark.marks = marks
				if mark.marks > mark.work.maxima:
					messages.error(request, "Mark out of maxima")
					form =  MarkForm()
				else:
					mark.save()
					page += 1
	try:
		pagination = paginator.page(page)
		student = pagination.object_list[0]
	except :
		return redirect(course_view, slug=work.course.slug)

	form =  MarkForm()
	return render(request, 'teacher_forms.html', locals())


@login_required
def student_number_view(request, work_id):
	work = get_object_or_404(Work, pk=work_id)
	student_number_form = StudentNumberForm(request.POST)
	if request.method == 'POST':
		if student_number_form.is_valid():
			student_number = student_number_form.cleaned_data['student_number']
			return redirect(random_add_mark_view, work_id=work_id, pk=student_number)

	student_number_form = StudentNumberForm()
	return render(request, 'teacher_forms.html', locals())

@login_required
def random_add_mark_view(request, work_id, pk):
	work = get_object_or_404(Work, pk=work_id)
	student = get_object_or_404(Student, Class = work.course.Class, pk=pk)
	random_add_mark_form = StudentMarkForm(request.POST)
	if request.method == 'POST':
		if random_add_mark_form.is_valid():
			marks = random_add_mark_form.cleaned_data['student_marks']
			
			try:
				mark = Mark.objects.get(student=student, work=work)
				mark.marks = marks
				mark.save()
			except:
				mark = Mark(student = student, work=work)
				mark.marks = marks
				mark.save()
			return redirect(student_number_view, work_id=work_id)

	random_add_mark_form = StudentMarkForm()
	return render(request, 'teacher_forms.html', locals())


@login_required
def load_mark(request, id):
	work = get_object_or_404(Work, id=id)
	# students = Student.objects.filter(Class = work.course.Class)
	if request.method == 'POST':
		load_form = LoadMarkForm(request.POST, request.FILES)
		if load_form.is_valid():
			wb = openpyxl.load_workbook(filename=BytesIO(request.FILES['fichier'].read()))
			sheet = wb[wb.sheetnames[0]]
			i = 1
			while 1:
				if not sheet.cell(column=1, row=i).value : break
				firstname = sheet.cell(column=1, row=i).value
				lastname = sheet.cell(column=2, row=i).value
				marks = sheet.cell(column=3, row=i).value
				student = get_object_or_404(Student, firstname=firstname, lastname=lastname, Class=work.course.Class)

				try:
					mark = Mark.objects.get(student=student, work=work)
					mark.marks = marks
					mark.save()
				except:
					Mark(student = student, work=work, marks=marks).save()

				i += 1
			return redirect(work_view, slug=work.course.slug, id=work.id)
	load_form = LoadMarkForm(request.FILES)
	return render(request, 'teacher_forms.html', locals())

@login_required
def modify_mark_view(request, student_id, id):
	show_hidden = "hidden"
	student = get_object_or_404(Student, id=student_id)
	classe = student.Class
	course = get_object_or_404(Course, Class=classe)
	work = get_object_or_404(Work, course=course, id=id)
	mark = get_object_or_404(Mark, student=student,work=work)
	modify_mark_form = MarkForm(request.POST, instance=mark)
	if request.method == 'POST':
		if modify_mark_form.is_valid:
			form = modify_mark_form.save(commit=False)
			if mark.marks > mark.work.maxima:
				messages.error(request, "Mark out of maxima")
				modify_mark_form = MarkForm(instance=mark)
			else:
				form.save()
				messages.success(request, "Mark modified successfuly")

			return redirect(work_view, slug=course.slug, id=work.id)
	modify_mark_form = MarkForm(instance=mark)
	return render(request, 'teacher_forms.html', locals())

@login_required
def view_mark_view(request):
	show_hidden = "hidden"
	marks = Mark.objects.all()
	return render(request, 'view_mark.html', locals())

@login_required
def course_view(request, slug):
	show_hidden = "hidden"
	cours = get_object_or_404(Course, slug=slug)
	classe = cours.Class
	students = Student.objects.filter(Class=classe)
	works = Work.objects.filter(course=cours)

	return render(request, "view_course.html", locals())

@login_required
def student_work_view(request, slug):
	show_hidden = "hidden"
	cours = get_object_or_404(Course, slug=slug)
	if request.method == 'POST':
		classe = cours.Class
		students = Student.objects.filter(Class=classe)

		for student in students:
			Mark(student=student, course=cours).save()

	marks = Mark.objects.filter(course=cours)


	return render(request, "view_mark.html", locals())

@login_required
def add_work_view(request, slug):
	show_hidden = "hidden"

	cours = get_object_or_404(Course, slug=slug)
	if request.method == "POST":
		classe = cours.Class
		add_work_form = WorkForm(request.POST or None)
		if add_work_form.is_valid():
			work = add_work_form.save(commit=False)
			work_number = Work.objects.filter(category=work.category).count()
			work.course = cours
			work.number = work_number + 1
			work.save()
			if True:
				messages.success(request, f"Work n° {work.number} of {work.category.category} added successfuly !")
			else:
				messages.error(request, 'Task Faild');

			return redirect(course_view, slug=slug)



	add_work_form = WorkForm()
	return render(request, 'teacher_forms.html', locals())



@login_required
def edit_work(request, id):
	show_hidden = "hidden"
	work = get_object_or_404(Work, id=id)
	cours = work.course
	classe = cours.Class
	school = classe.school
	teacher = get_object_or_404(Role, url='teacher')
	attribution = get_object_or_404(Attribution, user=request.user, school=school, role=teacher)
	edit_work_form = WorkForm(request.POST, instance=work)
	if request.method == "POST":
		if Attribution:
			if edit_work_form.is_valid():
				edit_work_form.save()
				messages.success(request, f"Work n° {work.number} of {work.category.category} modified successfuly !")
				return redirect(work_view, slug=cours.slug, id=work.id)
	edit_work_form = WorkForm(instance=work)
	return render(request, "teacher_forms.html", locals())








@login_required
def delete_work_view(request, id):
	show_hidden = "hidden"
	work = get_object_or_404(Work, id=id)
	cours = work.course
	classe = cours.Class
	school = classe.school
	teacher = get_object_or_404(Role, url='teacher')
	attribution = get_object_or_404(Attribution, user=request.user, school=school, role=teacher)
	if attribution:
		work.delete()
		messages.success(request, f"Work n° {work.number} of {work.category.category} deleted successfuly !")
		return redirect(course_view, slug=cours.slug)
	messages.error(request, 'Deleting error failed !')
	return redirect(work_view, cours.slug, work.id)


@login_required
def work_view(request, slug, id):
	show_hidden = "hidden"
	course = get_object_or_404(Course, slug=slug)
	work = get_object_or_404(Work, course = course, id=id )
	marks = Mark.objects.filter(work = work)
	# students = Student.objects.filter(Class=course.Class)
	return render(request, 'view_work.html', locals())


@login_required
def compile_view(request, classe_id):
	show_hidden = "hidden"
	classe = get_object_or_404(Class, id=classe_id)
	cours = Course.objects.filter(Class=classe)
	students = Student.objects.filter(Class=classe)
	for course in cours:
		for student in students:
			works = Work.objects.filter(course=course)
			for work in works:
				tj, exam, maxima = 0, 0, 0
				marks = Mark.objects.filter(work=work, student=student)
				for mark in marks:
					try:
						student_work = StudentWork.objects.get(course=course, student = student)
					except:
						student_work = StudentWork(course=course, student = student)

					if work.category.category == "1st Term":
						
						if (work.work_type.work_type == "Exam" and work.is_valid):
							exam += mark.marks
							maxima += mark.work.maxima
							student_work.exam_first_term = (exam*(mark.work.course.max_ponderation/2))/maxima

						elif (work.work_type.work_type == "Work" and work.is_valid):
							tj += mark.marks
							maxima += mark.work.maxima
							student_work.work_first_term = (tj*(mark.work.course.max_ponderation/2))/maxima

						student_work.save()

					if work.category.category == "2nd Term":
						if (work.work_type.work_type == "Exam" and work.is_valid):
							exam += mark.marks
							maxima += mark.work.maxima
							student_work.exam_second_term = (exam*(mark.work.course.max_ponderation/2))/maxima

						elif (work.work_type.work_type == "Work" and work.is_valid):
							tj += mark.marks
							maxima += mark.work.maxima
							student_work.work_second_term = (tj*(mark.work.course.max_ponderation/2))/maxima
						student_work.save()
							
					if work.category.category == "3rd Term":
						if (work.work_type.work_type == "Exam" and work.is_valid):
							maxima += mark.work.maxima
							exam += mark.marks
							student_work.exam_third_term = (exam*(mark.work.course.max_ponderation/2))/maxima

						elif (work.work_type.work_type == "Work" and work.is_valid):
							tj += mark.marks
							maxima += mark.work.maxima
							student_work.work_third_term = (tj*(mark.work.course.max_ponderation/2))/maxima

						student_work.save()


						
					

				tj=0
				exam=0
	students_works = StudentWork.objects.filter(student__in = students)


		

	return render(request, 'view_mark_student.html', locals())


