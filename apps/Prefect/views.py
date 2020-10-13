from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
# from crispy_forms.helper import FormHelper
# from apps.utils import PdfRender
from apps.Base.models import Attribution, Profil
from apps.Base.forms import InscriptionForm
from apps.Director.models import School
from apps.TeachersAndTitulars.models import Mark
from .forms import *

# Create your views here.

def isPrefectAt(request, school_id):
	school = get_object_or_404(School, id=school_id)
	prefect = get_object_or_404(Role, url="prefect")
	attribution = get_object_or_404(Attribution, \
		user=request.user, school=school, role=prefect)
	return True


@login_required
def index(request, school):
	school = get_object_or_404(School, slug=school)
	h3 = "prefecture - "+str(school)
	classes = Class.objects.filter(school=school)
	return render(request, "prefect_panel.html", locals())

@login_required
def bulettins(request, class_id):
	classe = get_object_or_404(Class, id=class_id)
	students = Student.objects.filter(Class=classe)
	marks = StudentWork.objects.filter(student__in=students)
	isPrefectAt(request, classe.school.id)
	# pdf = PdfRender("prefect_bulettins.html", locals()).render_to_pdf()
	# return HttpResponse(pdf, content_type='application/pdf')
	return render(request, "prefect_bulettins.html", locals())

@login_required
def bulettin(request, student_id, class_id):
	classe = get_object_or_404(Class, id=class_id)
	student = get_object_or_404(Student, id=student_id)
	marks = StudentWork.objects.filter(student=student)
	isPrefectAt(request, classe.school.id)
	# pdf = PdfRender("prefect_bulettin.html", locals()).render_to_pdf()
	# return HttpResponse(pdf, content_type='application/pdf')
	return render(request, "prefect_bulettin.html", locals())

@login_required
def courses(request, id_class):
	classe = get_object_or_404(Class, id=id_class)
	h3 = classe
	students = Student.objects.filter(Class=classe)
	isPrefectAt(request, classe.school.id)
	courses = Course.objects.filter(Class=classe)
	return render(request, "prefect_class.html", locals())

@login_required
def course(request, slug):
	cours = get_object_or_404(Course, slug=slug)
	h3 = cours
	# marks = Mark.objects.filter(course=cours)
	isPrefectAt(request, cours.Class.school.id)
	# courses = Course.objects.filter(Class=classe)
	return render(request, "prefect_cours.html", locals())

@login_required
def addClass(request, school):
	school = get_object_or_404(School, slug=school)
	isPrefectAt(request, school.id)
	class_form = ClassForm(request.POST or None, request.FILES)
	header_title = "Add New Class"
	if request.method == "POST":
		if class_form.is_valid():
			classe = class_form.save(commit=False)
			classe.school = school
			classe.save()
			messages.success(request, "Class added successfully")
			if 'quit' in request.POST:
				return redirect(index, school=school.slug)
	class_form = ClassForm()
	return render(request, "prefect_forms.html", locals())

@login_required
def modifyClass(request, class_id):
	classe = get_object_or_404(Class, id=class_id)
	isPrefectAt(request, classe.school.id)
	header_title = "Modify Class"
	if request.method == "POST":
		modify_class_form = ClassForm(request.POST, instance=classe)
		if modify_class_form.is_valid():
			modify_class_form.save()
			messages.success(request, "Class updated successfully")
			if 'quit' in request.POST:
				return redirect(courses, class_id)
	modify_class_form = ClassForm(instance=classe)
	return render(request, "prefect_forms.html", locals())

@login_required
def modifyCourse(request, slug):
	course = get_object_or_404(Course, slug=slug)
	isPrefectAt(request, course.Class.school.id)
	ids_profs = Attribution.objects.values_list("user", flat=True).filter(school=course.Class.school)
	profs = Profil.objects.filter(id__in=ids_profs)
	modify_course_form = CourseForm(profs, request.POST, instance=course)
	header_title = "Modify Course"
	if request.method == "POST":
		if modify_course_form.is_valid():
			course.save()
			messages.success(request, "Course updated successfully")
			if 'quit' in request.POST:
				return redirect(course, slug)
	modify_course_form = CourseForm(profs, instance=course)
	return render(request, "prefect_forms.html", locals())

@login_required
def addSection(request, school):
	school = get_object_or_404(School, slug=school)
	isPrefectAt(request, school.id)
	section_form = SectionForm(request.POST or None, request.FILES)
	header_title = "Section"
	if request.method == "POST":
		if section_form.is_valid():
			section_form.save()
			messages.success(request, "Section added successfully")
			if "add" not in request.POST:
				return redirect(index, school=school.slug)
	section_form = SectionForm()
	return render(request, "prefect_forms.html", locals())

# @login_required
# def addLevel(request, school):
# 	form = LevelForm(request.POST or None, request.FILES)
# 	if request.method == "POST":
# 		if form.is_valid():
# 			form.save()
# 			return redirect(index, school=school)
# 		form = LevelForm()
# 	return render(request, "prefect_forms.html", locals())

@login_required
def addCourse(request, classe_id):
	header_title = "Course"
	classe = get_object_or_404(Class, pk=classe_id)
	isPrefectAt(request, classe.school.id)
	ids_profs = Attribution.objects.values_list("user", flat=True).filter(school=classe.school)
	profs = Profil.objects.filter(id__in=ids_profs)
	add_course_form = CourseForm(profs, request.POST)
	if request.method == "POST":
		if add_course_form.is_valid():
			course = add_course_form.save(commit=False)
			course.Class = classe
			course.save()
			messages.success(request, "Course added successfully")
			if "quit" in request.POST:
				return redirect(courses, id_class=classe_id)
	add_course_form = CourseForm(profs)
	return render(request, "prefect_forms.html", locals())

@login_required
def addStudent(request, classe):
	classe = get_object_or_404(Class, id=classe)
	isPrefectAt(request, classe.school.id)
	student_form = StudentForm(request.POST)
	header_title = "Add New Student"
	if request.method == "POST":
		if student_form.is_valid():
			student = student_form.save(commit=False)
			student.Class = classe
			student.save()
			messages.success(request, "Student added successfully")
			if 'quit' in request.POST:
				return redirect(courses, id_class=classe.id)
	student_form = StudentForm()
	return render(request, "prefect_forms.html", locals())

@login_required
def loadStudent(request, classe):
	classe = get_object_or_404(Class, id=classe)
	isPrefectAt(request, classe.school.id)
	if request.method == 'POST':
		form = LoadStudentsForm(request.POST, request.FILES)
		if form.is_valid():
			handle_uploaded_file(classe, request.FILES['list_student'])
			return redirect(courses, id_class=classe.id)
			messages.success(request, "Students loaded successfully")

	form = LoadStudentsForm(request.FILES)
	return render(request, "prefect_forms.html", locals())

def handle_uploaded_file(classe, fichier):
	wb = openpyxl.load_workbook(filename=BytesIO(fichier.read()))
	sheet = wb[wb.sheetnames[0]]
	a, i = [], 1
	while 1:
		if not sheet.cell(column=1, row=i).value : break
		nom = sheet.cell(column=1, row=i).value
		prenom = sheet.cell(column=2, row=i).value
		Student(firstname=nom, lastname=prenom, Class=classe).save()
		i+=1


@login_required
def deleteStudent(request, student_id):
	student = get_object_or_404(Student, id=student_id)
	id_class = student.Class.id
	isPrefectAt(request, student.Class.school.id)
	student.delete()
	messages.success(request, "student deleted successfully")
	return redirect(courses, id_class=id_class)

@login_required
def editStudent(request, student_id):
	student = get_object_or_404(Student, id=student_id)
	isPrefectAt(request, student.Class.school.id)
	form = StudentForm(request.POST, instance=student)
	if request.method == "POST":
		if form.is_valid():
			form.save()
			messages.success(request, "Student updated successfully")
			if 'quit' in request.POST:
				return redirect(courses, id_class=student.Class.id)
	form = StudentForm(instance=student)
	return render(request, "prefect_forms.html", locals())

@login_required
def deleteClass(request, class_id):
	classe = get_object_or_404(Class, id=class_id)
	school = classe.school
	attribution = get_object_or_404(Attribution, user=request.user, school=school)
	if attribution.role.role in ("director", "prefect"):
		classe.delete()
		messages.success(request, "Class deleted successfully")
		return redirect(index, school=school.slug)
	return redirect(classe(request, classe))

@login_required
def deleteCourse(request, slug):
	cours = get_object_or_404(Course, slug=slug)
	classe = cours.Class
	school = classe.school
	attribution = get_object_or_404(Attribution, user=request.user, school=school)
	if attribution.role.role in ("director", "prefect"):
		cours.delete()
		messages.success(request, "Course deleted successfully")
		return redirect(courses, id_class=classe.id)
	return redirect(course, slug=cours.slug)
