from django.shortcuts import render, get_object_or_404, redirect, reverse
from django.http import HttpResponse, HttpResponseRedirect
from .forms import *
from .models import *
from apps.Base.models import *
from apps.Base.forms import *
from apps.Prefect.models import *
from apps.Director.models import *
from django.db.models import Q
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils.datastructures import MultiValueDictKeyError
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.files import File
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

# Create your views here.
@login_required
def index(request, school):
	school = get_object_or_404(School, slug=school)
	h3 = "Director section"
	show_hidden = "hidden"
	return render(request, "director_panel.html", locals())

@login_required
def school_details_view(request, slug):
	school = get_object_or_404(School, slug=slug)
	classes = Class.objects.filter(school=school)
	show_hidden = "hidden"
	return render(request, "school_details.html", locals())

@login_required
def view_employee_view(request, slug):
	school = get_object_or_404(School, slug = slug)
	classes = Class.objects.filter(school = school)
	font_image = "fa fa-eye"
	go_home = "home"
	heading_title = "View All Employees"
	h3 = "All employees of "+school.name+"."
	show_hidden = "hidden"

	connected_user = get_object_or_404(User, pk = request.user.id)
	rolee = get_object_or_404(Role, url = "director")
	check_user = Attribution.objects.filter(school = school.id, user = connected_user, role = rolee.id).count()
	all_attributions = Attribution.objects.all()

	if check_user == 1:
		all_attributes = Attribution.objects.filter(school=school.id)
	else:
		return redirect('../../../')

	return render(request, "view_employee_view.html", locals())

@login_required
def add_users_list(request, school, roley):
	school = get_object_or_404(School, slug=school)
	# roley = get_object_or_404(Role, id=roley)
	font_image = "fa fa-plus-square"
	heading_title = "Add employee"
	h3 = "Employees"
	h4 = "All users you create, they will have the same default password, 'password'"
	go_home = "home"
	show_hidden = "hidden"
	connected_user = get_object_or_404(User, pk=request.user.id)
	rolee = get_object_or_404(Role, url="director")

	check_user = Attribution.objects.filter(school=school.id, user=connected_user.id, role=rolee.id).count()
	all_attributions = Attribution.objects.all()
	search_role = Role.objects.all()

	if check_user == 1:
		if request.method == "POST":
			form = AddUsersForm(request.POST, request.FILES)
			if form.is_valid():
				try:
					j = load_users(school, roley, request.FILES['allusers_list'])
					if not j:
						messages.error(request, "List is already loaded")

					show_hidden = "show"
					home = "home"
					btn_class = "btn btn-lg btn-primary mt-5 btn-block"
					back = "Go back home"
					if j:messages.success(request, "List of employees is loaded successfully!")

					return redirect(view_employee_view,school.slug)
				except OSError:
					messages.error(request,"Format not supported!")
				except NameError:
					messages.error(request,"Format not supported!")

				# except BadZipFile:
				# 	messages.error(request,"Format not supported!")
				else:
					messages.error(request,"In list, some usernames is already used!")

		form = AddUsersForm(request.FILES)

	else:
		messages.danger(request, "Something went wrong!")
		return redirect(add_user_view,school)
	return render(request, 'add_users.html', locals())

def load_users(school, roley, fichier):
	wb = openpyxl.load_workbook(filename=BytesIO(fichier.read()))
	role = get_object_or_404(Role, id=roley)
	sheet = wb[wb.sheetnames[0]]
	i,j = 1,0
	while 1:
		if not sheet.cell(column=1, row=i).value : break
		fname = sheet.cell(column=1, row=i).value
		lname = sheet.cell(column=2, row=i).value
		password="password"
		try:
			user = User.objects.create_user(username = fname + lname, password = password)
			user.first_name = fname
			user.last_name = lname
			user.save()
			Profil.objects.get_or_create(user = user, avatar = File(open('media/avatars/thor.jpg', 'rb')))
			Attribution(user = user, school = school, role = role).save()
			j+=1
		except:
			print()
		i+=1

	return j

@login_required
def add_user_view(request, school):
	school = get_object_or_404(School, slug=school)
	font_image = "fa fa-plus-square"
	heading_title = "Add employee"
	h3 = "Employees"
	go_home = "home"
	h4 = "All users you create, they will have the same default password, 'password'"
	show_hidden = "hidden"
	connected_user = get_object_or_404(User, pk=request.user.id)
	rolee = get_object_or_404(Role, url="director")
	roley = get_object_or_404(Role, url="teacher")
	check_user = Attribution.objects.filter(school=school.id, user=connected_user.id, role=rolee.id).count()
	all_attributions = Attribution.objects.all()
	search_role = Role.objects.all()

	if check_user == 1:
		if request.method == "POST":
			form = AddUserForm(request.POST, request.FILES)
			if form.is_valid():
				try:
					username = form.cleaned_data['username']
					role = form.cleaned_data['role']
					password = "password"
					password2 = "password"
					if password==password2:
						user = User.objects.create_user(
							username=username,
							password=password)
						user.save()
						get_user = user.id
						get_username = user.username
						Attribution(user = user, role = role, school = school).save()
						p = Profil(user = user)
						p.avatar = File(open('media/avatars/thor.jpg', 'rb'))
						p.save()
						messages.success(request, username+" is is saved successfully!")
						show_hidden = "show"
						home = "home"
						btn_class = "btn btn-lg btn-primary mt-5 btn-block"
						back = "Go back home"
						return redirect(view_employee_view,school.slug)
				except:
					messages.error(request,"Username is already exist!")
		form = AddUserForm()
	else:
		return redirect('../../../')
	return render(request, 'add_user.html', locals())

@login_required
def home_view(request):
	font_image = "fa fa-plus-square"
	heading_title = "Hogi Admin Section"
	h3 = "Hogi Section"
	go_home = "home"
	total_schools = School.objects.all().count()
	total_provinces = Province.objects.all().count()
	total_communes = Commune.objects.all().count()
	total_zones = Zone.objects.all().count()
	total_schooltypes = SchoolType.objects.all().count()
	total_roles = Role.objects.all().count()
	show_hidden = "hidden"

	return render(request, "director_home.html", locals())

@login_required
def choose_user_view(request,school):
	school = get_object_or_404(School, slug = school)
	all_users1 = 0
	font_image = "fa fa-plus-square"
	heading_title = "Search user"
	h3 = "Searching..."
	show_hidden = "hidden"
	go_home = "home"

	rolee = get_object_or_404(Role, url = "director")
	check_user = Attribution.objects.filter(school = school.id, user = request.user.id, role = rolee.id).count()

	if check_user == 1:
		form = SearchUserForm(request.POST or None, request.FILES)
		if request.method == "POST":
			if form.is_valid():
				get_searched_user = form.cleaned_data['search_user']
				user_name = User.objects.filter(
					Q(username__startswith=get_searched_user) |
					Q(username__endswith=get_searched_user) |
					Q(username__icontains=get_searched_user)
					)
				all_users1 = user_name.count()

				if all_users1 > 0 and all_users1 == 1:
					form = SearchUserForm()
					msg = "User Found!"

				elif all_users1 >= 2:
					form = SearchUserForm()

				else:
					msg = "User not Found!"
	else:
		return redirect('../../../')

	return render(request, "choose_user.html", locals())

##############################################################
#####################      Addition     ######################
##############################################################

@login_required
def attribute_view(request, id, slug):
	font_image = "fa fa-plus-square"
	heading_title = "Attribute Role"
	show_hidden = "hidden"
	go_home = "home"

	get_user_id = get_object_or_404( User, pk = id )
	get_school_slug = get_object_or_404(School, slug = slug)
	all_attributes = Attribution.objects.filter(school = get_school_slug.id)
	form = AttributionForm(request.POST or None, request.FILES)
	h3 = "Attributions."
	h2 = get_school_slug

	rolee = get_object_or_404(Role, url = "director")
	check_user = Attribution.objects.filter(school = get_school_slug.id, user = request.user.id, role = rolee.id).count()

	if check_user == 1:
		if request.method == "POST":
			if form.is_valid():
				try:
					get_user_id1 = get_user_id.id
					role = form.cleaned_data['role']
					Attribution(user = get_user_id, role = role, school = get_school_slug ).save()
					show_hidden = "show"
					messages.success(request, role.role+" role is given to "+get_user_id.username+" successfully!")
				except:
					messages.error(request,role.role+" role is already given to "+get_user_id.username)
			form = AttributionForm()
	else:
		return redirect(attribute_view,get_school_slug)

	return render(request, "attribute_roles.html", locals())

@login_required
def role_view(request):
	font_image = "fa fa-plus-square"
	heading_title = "Add Role"
	h3 = "Roles"
	go_homee = "hpanel"
	all_roles = Role.objects.all()
	show_hidden = "hidden"
	form = RoleForm(request.POST or None, request.FILES)

	if request.method == "POST":
		if form.is_valid():
			form.save()
			show_hidden = "show"
			messages.success(request, form.cleaned_data['role']+" role is saved successfully!")
		form = RoleForm()
	return render(request, "add_role.html", locals())

@login_required
def add_province_view(request):
	font_image = "fa fa-plus-square"
	heading_title = "Add New Province"
	h3 = "Provinces"
	go_homee = "hpanel"
	all_provinces = Province.objects.all()
	show_hidden = "hidden"
	back_home = "hpanel"
	form = ProvinceForm(request.POST or None, request.FILES)
	if request.method == "POST":
		if form.is_valid():
			form.save()
			show_hidden = "show"
			messages.success(request, form.cleaned_data['name']+" province is saved successfully!")
		form = ProvinceForm()
	return render(request, "add_province.html", locals())

@login_required
def add_commune_view(request):
	font_image = "fa fa-plus-square"
	heading_title = "Add New Commune"
	h3 = "Communes"
	go_homee = "hpanel"
	all_communes = Commune.objects.all()
	show_hidden = "hidden"
	form = CommuneForm(request.POST or None, request.FILES)
	if request.method == "POST":
		if form.is_valid():
			form.save()
			show_hidden = "show"
			messages.success(request, form.cleaned_data['name']+" commune is saved successfully!")
		form = CommuneForm()
	return render(request, 'add_commune.html', locals())

@login_required
def add_zone_view(request):
	font_image = "fa fa-plus-square"
	heading_title = "Add New Zone"
	h3 = "Zones"
	go_homee = "hpanel"
	all_zones = Zone.objects.all()
	show_hidden = "hidden"
	form = ZoneForm(request.POST or None, request.FILES)
	if request.method == "POST":
		if form.is_valid():
			form.save()
			show_hidden = "show"
			messages.success(request, form.cleaned_data['name']+" zone is saved successfully!")
		form = ZoneForm()
	return render(request, "add_zone.html", locals())

@login_required
def add_schooltype_view(request):
	font_image = "fa fa-plus-square"
	heading_title = "Add New School Type"
	h3 = "School Type"
	go_homee = "hpanel"
	all_schooltypes = SchoolType.objects.all()
	show_hidden = "hidden"
	form = SchoolTypeForm(request.POST or None, request.FILES)
	if request.method == "POST":
		if form.is_valid():
			form.save()
			show_hidden = "show"
			messages.success(request, form.cleaned_data['name']+" school type is saved successfully!")
		form = SchoolTypeForm()
	return render(request, "add_schooltype.html", locals())

@login_required
def add_school_view(request):
	font_image = "fa fa-plus-square"
	heading_title = "Add New School"
	h3 = "Schools"
	go_homee = "hpanel"
	all_schools = School.objects.all()
	show_hidden = "hidden"
	rolee = get_object_or_404(Role,url = "director")
	form = SchoolForm(request.POST or None, request.FILES)

	if request.method == "POST":
		if form.is_valid():
			current_user = request.user
			name = form.cleaned_data['name']
			province = form.cleaned_data['province']
			commune = form.cleaned_data['commune']
			zone = form.cleaned_data['zone']
			school_type = form.cleaned_data['school_type']
			description = form.cleaned_data['description']

			save_school = School(added_by = current_user, name = name, province = province, commune = commune,
			zone = zone, school_type = school_type, description = description)
			save_school.save()
			Attribution(user = request.user, role = rolee, school = save_school).save()
			show_hidden = "show"
			messages.success(request, form.cleaned_data['name']+" school is saved successfully!")

		form = SchoolForm()
	return render(request, "add_school.html", locals())

##############################################################
#####################       Updates     ######################
##############################################################

#commune
@login_required
def update_commune_view(request,id):
	font_image = "fa fa-refresh"
	heading_title = "Update Commune"
	h3 = "Communes"
	go_homee = "hpanel"
	show_hidden = "hidden"
	selected_commune = get_object_or_404( Commune, pk = id )
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			form = CommuneForm(request.POST or None, request.FILES or None, instance = selected_commune)
			if form.is_valid():
				form.save()
				show_hidden = "show"
				messages.success(request, "Commune is updated successfully!")
				return redirect(home_view)
		else:
			data = {'id'           : selected_commune.id,
                    'name'         : selected_commune.name,
                    'province'     : selected_commune.province,
                    'slug'         : selected_commune.slug,
					'date'		   : selected_commune.date,
                    }
			form = CommuneForm( initial = data )
	return render( request, "update_commune.html", locals() )

#zone
@login_required
def update_zone_view(request,id):
	font_image = "fa fa-refresh"
	heading_title = "Add Zone"
	h3 = "Zones"
	go_homee = "hpanel"
	show_hidden = "hidden"
	selected_zone = get_object_or_404( Zone, pk = id )
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			form = ZoneForm(request.POST or None, request.FILES or None, instance = selected_zone)
			if form.is_valid():
				form.save()
				show_hidden = "show"
				messages.success(request, "Zone is updated successfully!")
				return redirect(home_view)
		else:
			data = {'id'           : selected_zone.id,
                    'name'         : selected_zone.name,
                    'commune'     : selected_zone.commune,
                    'slug'         : selected_zone.slug,
					'date'		   : selected_zone.date,
                    }
			form = ZoneForm( initial = data )
	return render( request, "update_zone.html", locals() )

#province
@login_required
def update_province_view(request,id):
	font_image = "fa fa-refresh"
	heading_title = "Update Province"
	h3 = "Provinces"
	go_homee = "hpanel"
	show_hidden = "hidden"
	back_home = "hpanel"
	selected_province = get_object_or_404( Province, pk = id )
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			form = ProvinceForm(request.POST or None, request.FILES or None, instance = selected_province)
			if form.is_valid():
				form.save()
				show_hidden = "show"
				messages.success(request, "Province is updated successfully!")

				return redirect(home_view)
		else:
			data = {'id'           : selected_province.id,
                    'name'         : selected_province.name,
                    'slug'         : selected_province.slug,
					'date'		   : selected_province.date,
                    }
			form = ProvinceForm( initial = data )
	return render( request, "update_province.html", locals() )

#school
@login_required
def update_school_view(request,id):
	font_image = "fa fa-refresh"
	heading_title = "Update School"
	h3 = "Schools"
	go_homee = "hpanel"
	show_hidden = "hidden"
	selected_school = get_object_or_404( School, pk = id )
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			form = SchoolForm(request.POST or None, request.FILES or None, instance = selected_school)
			if form.is_valid():
				form.save()
				show_hidden = "show"
				messages.success(request, "School is updated successfully!")
				return redirect(home_view)
		else:
			data = {'id'           : selected_school.id,
                    'name'         : selected_school.name,
                    'added_by'     : selected_school.added_by,
                    'zone'         : selected_school.zone,
                    'province'     : selected_school.province,
                    'commune'      : selected_school.commune,
                    'description'  : selected_school.description,
                    'school_type'  : selected_school.school_type,
                    'visible'      : selected_school.visible,
                    'slug'         : selected_school.slug,
					'date'		   : selected_school.date,
                    }
			form = SchoolForm( initial = data )
	return render( request, "update_school.html", locals() )

#school type
@login_required
def update_schooltype_view(request,id):
	font_image = "fa fa-refresh"
	heading_title = "Update School Type"
	h3 = "School Type"
	go_homee = "hpanel"
	show_hidden = "hidden"
	selected_schooltype = get_object_or_404( SchoolType, pk = id )
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			form = SchoolTypeForm(request.POST or None, request.FILES or None, instance = selected_schooltype)
			if form.is_valid():
				form.save()
				show_hidden = "show"
				messages.success(request, "School type is updated successfully!")
				return redirect(home_view)
		else:
			data = {'id'                : selected_schooltype.id,
                    'education'         : selected_schooltype.education,
                    }
			form = SchoolTypeForm( initial = data )
	return render( request, "update_schooltype.html", locals() )

#role
@login_required
def update_role_view(request,id):
	font_image = "fa fa-refresh"
	heading_title = "Update Role"
	h3 = "Roles"
	go_homee = "hpanel"
	show_hidden = "hidden"
	selected_role = get_object_or_404( Role, pk = id )
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			form = RoleForm(request.POST or None, request.FILES or None, instance = selected_role)
			if form.is_valid():
				form.save()
				show_hidden = "show"
				messages.success(request, "Role is updated successfully!")
				return redirect(home_view)
		else:
			data = {'id'          : selected_role.id,
                    'role'        : selected_role.role,
					'url'         : selected_role.url,
                    }
			form = RoleForm( initial = data )
	return render( request, "update_role.html", locals() )

#attribution
@login_required
def update_attribution_view(request,id,slug):
	font_image = "fa fa-refresh"
	heading_title = "Update Attribution"
	h3 = "Attributions"
	go_home = "home"
	show_hidden = "hidden"
	selected_attribution = get_object_or_404( Attribution, pk = id )
	school = get_object_or_404(School,slug = slug)

	connected_user = get_object_or_404(User, pk = request.user.id)
	rolee = get_object_or_404(Role, url = "director")
	check_user = Attribution.objects.filter(school = school.id, user = connected_user.id, role = rolee.id).count()
	all_attributions = Attribution.objects.all()

	if not request.user.is_authenticated:
		return redirect( connexion )
	elif check_user == 1:
		all_attributes = Attribution.objects.filter(school = school.id)
		if request.method == "POST":
			form = AttributionForm(request.POST or None, request.FILES or None, instance = selected_attribution)
			if form.is_valid():
				form.save()
				show_hidden = "show"
				messages.success(request, "Attribution role is updated successfully!")
				return redirect(view_employee_view,slug)
		else:
			data = {'id'       	: selected_attribution.id,
                    'user' 		: selected_attribution.user,
					'role'		: selected_attribution.role,
					'school'	: selected_attribution.school,
                    }
			form = AttributionForm( initial = data )
	else:
		return redirect('../../../../')
	return render( request, "update_attribution.html", locals() )

###############################################################
#####################       Deletion     ######################
###############################################################

#commune
@login_required
def delete_commune_view(request,slug):
	font_image = "fa fa-trash"
	heading_title = "Delete Commune"
	h3 = "Communes"
	go_homee = "hpanel"
	show_hidden = "hidden"
	selected_commune = get_object_or_404( Commune, slug = slug )
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			selected_commune.delete()
			show_hidden = "show"
			messages.success(request, "Commune is deleted successfully!")
			return redirect(home_view)
		return render( request, 'delete_commune.html', locals() )

#zone
@login_required
def delete_zone_view(request,slug):
	font_image = "fa fa-trash"
	heading_title = "Delete Zone"
	h3 = "Zones"
	go_homee = "hpanel"
	show_hidden = "hidden"
	selected_zone = get_object_or_404( Zone, slug = slug )
	h4 = "You are going to delete "+selected_zone.name+" zone."
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			selected_zone.delete()
			show_hidden = "show"
			messages.success(request, "Zone is deleted successfully!")
			return redirect(home_view)

	return render( request, 'delete_zone.html', locals() )

#province
@login_required
def delete_province_view(request,slug):
	font_image = "fa fa-trash"
	heading_title = "Delete Province"
	h3 = "Provinces"
	go_homee = "hpanel"
	show_hidden = "hidden"
	selected_province = get_object_or_404( Province, slug = slug )
	h4 = "You are going to delete '"+selected_province.name+"' province."
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			selected_province.delete()
			show_hidden = "show"
			messages.success(request, "Province is deleted successfully!")
			return redirect(home_view)
	return render( request, 'delete_province.html', locals() )

#school
@login_required
def delete_school_view(request,slug):
	font_image = "fa fa-trash"
	heading_title = "Delete School"
	h3 = "Schools"
	go_homee = "hpanel"
	show_hidden = "hidden"
	selected_school = get_object_or_404( School, slug = slug )
	h4 = "You are going to delete '"+selected_school.name+"' school."
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			selected_school.delete()
			show_hidden = "show"
			messages.success(request, "School is deleted successfully!")
			return redirect(home_view)
	return render( request, 'delete_school.html', locals() )

#school type
@login_required
def delete_schooltype_view(request,id):
	font_image = "fa fa-trash"
	heading_title = "Delete School Type"
	h3 = "School Type"
	go_homee = "hpanel"
	show_hidden = "hidden"
	selected_schooltype = get_object_or_404( SchoolType, pk = id )
	h4 = "You are going to delete '"+selected_schooltype.education+"'"
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			selected_schooltype.delete()
			show_hidden = "show"
			messages.success(request, "School type is deleted successfully!")
			return redirect(home_view)
	return render( request, 'delete_schooltype.html', locals() )

#school type
@login_required
def delete_role_view(request,id):
	font_image = "fa fa-trash"
	heading_title = "Delete Role"
	h3 = "Roles"
	go_homee = "hpanel"
	show_hidden = "hidden"
	selected_role = get_object_or_404( Role, pk = id )
	h4 = "You are going to delete '"+selected_role.role+"' role."
	if not request.user.is_authenticated:
		return redirect( connexion )
	else:
		if request.method == "POST":
			selected_role.delete()
			show_hidden = "show"
			messages.success(request, "Role is deleted successfully!")
			return redirect(home_view)

	return render( request, 'delete_role.html', locals() )

#attribution
@login_required
def delete_attribution_view(request,id,slug):
	font_image = "fa fa-trash"
	heading_title = "Delete Attribution"
	h3 = "Attributions"
	go_home = "home"
	show_hidden = "hidden"
	selected_attribution = get_object_or_404( Attribution, pk = id )
	school = get_object_or_404(School, slug = slug )
	h4 = selected_attribution.user.username+" will no longer be a "+selected_attribution.role.role+" at "+selected_attribution.school.name+"."

	connected_user = get_object_or_404(User, pk = request.user.id)
	rolee = get_object_or_404(Role, url="director")
	check_user = Attribution.objects.filter(school = school.id, user = connected_user.id, role = rolee.id).count()
	all_attributions = Attribution.objects.all()

	if not request.user.is_authenticated:
		return redirect( connexion )
	elif check_user == 1:
		if request.method == "POST":
			selected_attribution.delete()
			show_hidden = "show"
			messages.success(request, "Attribution is deleted successfully!")
			return redirect(view_employee_view,slug)
	else:
		return redirect(home_view)

	return render( request, 'delete_attribution.html', locals() )
